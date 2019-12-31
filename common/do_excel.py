# -*- coding: utf-8 -*-
# @Time    : 2019/11/6 14:41
# @Software: PyCharm Community Edition
# @Author  : Ada
# @File    : do_excel.py
# 完成excel的读与写
import openpyxl

from API.API_7.common import http_request

"""
测试用例类，每个测试用例，实际上就是它的一个实例
"""
class Case:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.url=None
        self.data=None
        self.method=None
        self.excepted=None
        self.actual=None
        self.result=None
        self.sql=None
class DoExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]
    def get_cases(self):
        max_row=self.sheet.max_row#获取最大行数

        cases=[]#列表，存放所有的测试用例
        for r in range(2,max_row+1):

            case=Case()#实例
            case.case_id=self.sheet.cell(row=r,column=1).value
            case.title=self.sheet.cell(row=r, column=2).value
            case.url=self.sheet.cell(row=r, column=3).value
            case.data=self.sheet.cell(row=r, column=4).value
            case.method=self.sheet.cell(row=r, column=5).value
            case.expected=self.sheet.cell(row=r, column=6).value
            case.sql=self.sheet.cell(row=r,column=9).value#执行的sql
            cases.append(case)

        self.workbook.close()
        return cases#返回case列表

    def write_result(self,row,actual,result):
        sheet=self.workbook[self.sheet_name]
        sheet.cell(row,7).value= actual
        sheet.cell(row,8).value = result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()

if __name__=="__main__":
    from API.API_3.common import contants
    do_excel=DoExcel(contants.case_file,sheet_name='login')
    cases=do_excel.get_cases()
    http_request= http_request.HTTPRequest()
    for case in cases:
        # print(case.case_id)
        # print(case.title)
        # print(case.url)
        # print(case.data)
        # print(case.method)
        # print(case.excepted)
        print(case.__dict__)
        print(type(case.data))
        resp=http_request.request(case.method,case.url,case.data)
        print(resp.status_code)
        print(resp.text)#响应文本
        resp_dict = resp.json()#返回字典
        print(resp_dict)

        actual=resp.text
        if case.excepted == actual:#判断期望结果和实际结果是否一致
            do_excel.write_result(case.case_id+1,actual,'PASS')

        else:
            do_excel.write_result(case.case_id+1,actual,'FAIL')